#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
核心算法模块：发票-销售数据合并
从 merge_all_sheets.py 提取，参数化为可导入的模块
"""

import pandas as pd
import shutil
import tempfile
import os
from difflib import SequenceMatcher
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 默认Sheet配置：(sheet名称, 开票数量列名, 开票金额列名, 允许匹配的税率列表)
DEFAULT_SHEET_CONFIGS = [
    ('蔬菜', '开票数量', '开票金额', ['13%', '免税']),
    ('肉蛋', '开票数量', '开票金额', ['9%', '13%', '免税']),
    ('9%', '已开票数量', '已开票金额', ['9%']),
    ('13%', '已开票数量', '已开票金额', ['13%']),
]


def similarity_ratio(str1, str2):
    """计算两个字符串的相似度（0-1之间）"""
    return SequenceMatcher(None, str1, str2).ratio()


def find_best_match(target, candidates, threshold=0.75):
    """在候选列表中找到与目标字符串最匹配的项"""
    best_match = None
    best_ratio = 0

    for candidate in candidates:
        ratio = similarity_ratio(target, candidate)
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = candidate

    if best_ratio >= threshold:
        return best_match, best_ratio
    return None, 0


def clean_product_name(name):
    """清洗商品名称：去掉前后空格、统一括号等"""
    if pd.isna(name):
        return name
    name_str = str(name).strip()
    name_str = name_str.replace('(', '(').replace(')', ')')
    name_str = name_str.replace('【', '[').replace('】', ']')
    name_str = ' '.join(name_str.split())
    return name_str


def remove_category_prefix(name):
    """去掉商品名称前的*分类*前缀"""
    if pd.isna(name):
        return name
    name_str = str(name)
    if name_str.startswith('*'):
        parts = name_str.split('*', 2)
        if len(parts) >= 3:
            return parts[2]
    return name_str


def collect_all_matches(df_sales, invoice_dict, invoice_product_names,
                        sheet_name, qty_col, amount_col, allowed_tax_rates, threshold=0.75):
    """收集单个sheet的所有可能匹配"""
    matches = []

    for idx, row in df_sales.iterrows():
        product_name = row['商品名称']

        # 尝试完全匹配
        matched = False
        for tax_rate in allowed_tax_rates:
            key = (product_name, tax_rate)
            if key in invoice_dict:
                matches.append({
                    'sheet_name': sheet_name,
                    'row_idx': idx,
                    'sales_product': product_name,
                    'invoice_key': key,
                    'invoice_product': product_name,
                    'tax_rate': tax_rate,
                    'similarity': 1.0,
                    'qty_col': qty_col,
                    'amount_col': amount_col,
                    'is_fuzzy': False
                })
                matched = True
                break

        if not matched:
            # 尝试模糊匹配
            best_match_product = None
            best_match_key = None
            best_ratio = 0

            for tax_rate in allowed_tax_rates:
                tax_rate_products = [name for (name, tr) in invoice_product_names if tr == tax_rate]
                if tax_rate_products:
                    match, ratio = find_best_match(product_name, tax_rate_products, threshold=threshold)
                    if match and ratio > best_ratio:
                        best_ratio = ratio
                        best_match_product = match
                        best_match_key = (match, tax_rate)

            if best_match_product:
                matches.append({
                    'sheet_name': sheet_name,
                    'row_idx': idx,
                    'sales_product': product_name,
                    'invoice_key': best_match_key,
                    'invoice_product': best_match_product,
                    'tax_rate': best_match_key[1],
                    'similarity': best_ratio,
                    'qty_col': qty_col,
                    'amount_col': amount_col,
                    'is_fuzzy': True
                })

    return matches


def deduplicate_matches(all_matches):
    """去重：对每个发票商品+税率组合，只保留相似度最高的匹配"""
    invoice_groups = {}
    for match in all_matches:
        invoice_key = match['invoice_key']
        if invoice_key not in invoice_groups:
            invoice_groups[invoice_key] = []
        invoice_groups[invoice_key].append(match)

    best_matches = []
    for invoice_key, matches in invoice_groups.items():
        best_match = max(matches, key=lambda m: m['similarity'])
        best_matches.append(best_match)

    return best_matches


def apply_matches(sheet_dataframes, best_matches, invoice_dict):
    """根据去重后的匹配关系，更新各个sheet的数据"""
    updated_dfs = {}
    sheet_stats = {}
    sheet_fuzzy_rows = {}
    all_fuzzy_matches = []
    matched_invoice_products = set()

    matches_by_sheet = {}
    for match in best_matches:
        sheet_name = match['sheet_name']
        if sheet_name not in matches_by_sheet:
            matches_by_sheet[sheet_name] = []
        matches_by_sheet[sheet_name].append(match)

    for sheet_name, df in sheet_dataframes.items():
        exact_match_count = 0
        fuzzy_match_count = 0
        fuzzy_rows = set()

        if sheet_name in matches_by_sheet:
            for match in matches_by_sheet[sheet_name]:
                idx = match['row_idx']
                invoice_key = match['invoice_key']
                invoice_product = match['invoice_product']
                qty_col = match['qty_col']
                amount_col = match['amount_col']

                df.at[idx, qty_col] = invoice_dict[invoice_key]['数量']
                df.at[idx, amount_col] = invoice_dict[invoice_key]['金额']

                matched_invoice_products.add(invoice_key)

                if match['is_fuzzy']:
                    fuzzy_match_count += 1
                    fuzzy_rows.add(idx + 2)

                    all_fuzzy_matches.append({
                        'Sheet': sheet_name,
                        '农品达商品名称': match['sales_product'],
                        '发票商品名称': invoice_product,
                        '税率': match['tax_rate'],
                        '相似度': f'{match["similarity"]:.2%}',
                        '数量': invoice_dict[invoice_key]['数量'],
                        '金额': invoice_dict[invoice_key]['金额']
                    })
                else:
                    exact_match_count += 1

        updated_dfs[sheet_name] = df
        sheet_stats[sheet_name] = {
            'exact_match': exact_match_count,
            'fuzzy_match': fuzzy_match_count,
            'unmatched': len(df) - exact_match_count - fuzzy_match_count,
            'total': len(df)
        }
        sheet_fuzzy_rows[sheet_name] = fuzzy_rows

    return updated_dfs, sheet_stats, sheet_fuzzy_rows, all_fuzzy_matches, matched_invoice_products


def process_merge(invoice_path, sales_path, output_dir, threshold=0.75,
                  sheet_configs=None, progress_callback=None):
    """
    主处理函数

    参数:
    - invoice_path: 发票文件路径
    - sales_path: 销售文件路径
    - output_dir: 输出目录路径
    - threshold: 模糊匹配阈值（默认0.75）
    - sheet_configs: Sheet配置列表，默认使用 DEFAULT_SHEET_CONFIGS
    - progress_callback: 进度回调函数 callback(step, total_steps, message)

    返回: {
        'output_file': str,
        'fuzzy_file': str 或 None,
        'unmatched_file': str 或 None,
        'stats': {
            'exact_matches': int,
            'fuzzy_matches': int,
            'unmatched': int,
            'sheet_stats': dict
        }
    }
    """
    if sheet_configs is None:
        sheet_configs = DEFAULT_SHEET_CONFIGS

    def progress(step, total, msg):
        if progress_callback:
            progress_callback(step, total, msg)

    total_steps = 7
    os.makedirs(output_dir, exist_ok=True)

    # 步骤1：读取发票数据
    progress(1, total_steps, "正在读取发票文件...")
    df_invoice = pd.read_excel(invoice_path)

    # 步骤2：聚合发票数据
    progress(2, total_steps, "正在聚合发票数据...")
    df_invoice['商品名称_清洗'] = df_invoice['货物或应税劳务名称'].apply(
        lambda x: clean_product_name(remove_category_prefix(x))
    )

    invoice_grouped = df_invoice.groupby(['商品名称_清洗', '税率']).agg({
        '数量': 'sum',
        '金额': 'sum'
    }).reset_index()

    invoice_dict = {}
    for _, row in invoice_grouped.iterrows():
        product_name = row['商品名称_清洗']
        tax_rate = row['税率']
        key = (product_name, tax_rate)
        invoice_dict[key] = {
            '数量': row['数量'],
            '金额': row['金额'],
            '商品名称': product_name
        }

    invoice_product_names = list(invoice_dict.keys())

    # 步骤3：复制原文件
    progress(3, total_steps, "正在准备输出文件...")
    sales_basename = os.path.splitext(os.path.basename(sales_path))[0]
    output_file = os.path.join(output_dir, f"{sales_basename}_已更新.xlsx")
    shutil.copy(sales_path, output_file)

    # 步骤4：收集所有可能的匹配
    progress(4, total_steps, "正在匹配商品名称...")
    all_possible_matches = []
    sheet_dataframes = {}

    for sheet_name, qty_col, amount_col, allowed_tax_rates in sheet_configs:
        try:
            df_sales = pd.read_excel(sales_path, sheet_name=sheet_name)
        except ValueError:
            continue

        # 清洗销售商品名称
        df_sales['商品名称'] = df_sales['商品名称'].apply(clean_product_name)

        df_sales[qty_col] = None
        df_sales[amount_col] = None
        sheet_dataframes[sheet_name] = df_sales

        matches = collect_all_matches(
            df_sales, invoice_dict, invoice_product_names,
            sheet_name, qty_col, amount_col, allowed_tax_rates,
            threshold=threshold
        )
        all_possible_matches.extend(matches)

    # 步骤5：去重
    progress(5, total_steps, "正在去重优化匹配结果...")
    best_matches = deduplicate_matches(all_possible_matches)

    # 步骤6：应用匹配
    progress(6, total_steps, "正在应用匹配结果...")
    updated_dfs, all_stats, sheet_fuzzy_rows, all_fuzzy_matches, matched_invoice_products = \
        apply_matches(sheet_dataframes, best_matches, invoice_dict)

    # 保存更新后的数据到Excel
    wb = load_workbook(output_file)
    red_font = Font(color="FF0000", bold=False)

    for sheet_name, qty_col, amount_col, _ in sheet_configs:
        if sheet_name not in updated_dfs:
            continue
        df_sales = updated_dfs[sheet_name]
        fuzzy_match_rows = sheet_fuzzy_rows[sheet_name]

        ws = wb[sheet_name]
        ws.delete_rows(2, ws.max_row)

        for r_idx, row in enumerate(dataframe_to_rows(df_sales, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx in fuzzy_match_rows:
                    cell.font = red_font

    wb.save(output_file)

    # 保存模糊匹配详情
    fuzzy_file = None
    if all_fuzzy_matches:
        fuzzy_file = os.path.join(output_dir, "模糊匹配详情.xlsx")
        df_fuzzy = pd.DataFrame(all_fuzzy_matches)
        df_fuzzy.to_excel(fuzzy_file, index=False, engine='openpyxl')

    # 保存未匹配记录
    unmatched_file = None
    all_invoice_keys = set(invoice_dict.keys())
    unmatched_invoice_keys = all_invoice_keys - matched_invoice_products

    if unmatched_invoice_keys:
        unmatched_rows = []
        for idx, row in df_invoice.iterrows():
            key = (row['商品名称_清洗'], row['税率'])
            if key in unmatched_invoice_keys:
                unmatched_rows.append(idx)

        df_unmatched = df_invoice.loc[unmatched_rows].copy()
        unmatched_file = os.path.join(output_dir, "未匹配记录.xlsx")

        wb_unmatched = Workbook()
        ws_unmatched = wb_unmatched.active
        ws_unmatched.title = "未匹配记录"

        original_columns = [col for col in df_unmatched.columns if col != '商品名称_清洗']
        for c_idx, col_name in enumerate(original_columns, 1):
            cell = ws_unmatched.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True)

        red_font_unmatched = Font(color="FF0000")
        for r_idx, (_, row) in enumerate(df_unmatched.iterrows(), 2):
            for c_idx, col_name in enumerate(original_columns, 1):
                value = row[col_name]
                cell = ws_unmatched.cell(row=r_idx, column=c_idx, value=value)
                cell.font = red_font_unmatched

        wb_unmatched.save(unmatched_file)

    progress(7, total_steps, "处理完成！")

    # 汇总统计
    total_exact = sum(s['exact_match'] for s in all_stats.values())
    total_fuzzy = sum(s['fuzzy_match'] for s in all_stats.values())
    total_unmatched_sales = sum(s['unmatched'] for s in all_stats.values())

    return {
        'output_file': output_file,
        'fuzzy_file': fuzzy_file,
        'unmatched_file': unmatched_file,
        'stats': {
            'exact_matches': total_exact,
            'fuzzy_matches': total_fuzzy,
            'unmatched_sales': total_unmatched_sales,
            'unmatched_invoices': len(unmatched_invoice_keys),
            'sheet_stats': all_stats
        }
    }
